from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import secrets
import string
import hashlib
import json
import tempfile
import os
from datetime import timedelta
from dotenv import load_dotenv
from api_routes import api
from ai_timetable import TimetableGenerator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)

jwt = JWTManager(app)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Register existing Blueprint
app.register_blueprint(api)

# Initialize the database
def init_db():
    conn = sqlite3.connect('timetable.db')
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK (role IN ('main_admin', 'dept_admin', 'staff')),
            department_id INTEGER,
            staff_role TEXT CHECK (staff_role IN ('assistant_professor', 'professor', 'hod')),
            subjects_selected TEXT,
            subjects_locked BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            code TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            code TEXT NOT NULL,
            department_id INTEGER NOT NULL,
            credits INTEGER DEFAULT 3,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS classrooms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            capacity INTEGER NOT NULL,
            department_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timetables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            day TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            subject_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            classroom_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (subject_id) REFERENCES subjects (id),
            FOREIGN KEY (staff_id) REFERENCES users (id),
            FOREIGN KEY (classroom_id) REFERENCES classrooms (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS constraints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER,
            role TEXT NOT NULL CHECK (role IN ('assistant_professor', 'professor', 'hod')),
            subject_type TEXT NOT NULL CHECK (subject_type IN ('theory', 'lab', 'both')),
            max_subjects INTEGER NOT NULL DEFAULT 1,
            max_hours INTEGER NOT NULL DEFAULT 8,
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Enhanced tables for new features
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS staff_registration_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            employee_id TEXT NOT NULL,
            email TEXT NOT NULL,
            department_id INTEGER NOT NULL,
            staff_role TEXT NOT NULL,
            contact_number TEXT,
            requested_by INTEGER NOT NULL,
            status TEXT DEFAULT 'pending' CHECK (status IN ('pending', 'approved', 'rejected')),
            approved_by INTEGER,
            credentials_generated BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (requested_by) REFERENCES users (id),
            FOREIGN KEY (approved_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS enhanced_constraints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            max_subjects INTEGER NOT NULL DEFAULT 1,
            max_hours_per_week INTEGER NOT NULL DEFAULT 8,
            subject_types TEXT,
            lab_faculty_required INTEGER DEFAULT 1,
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subject_choice_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            open_date TIMESTAMP NOT NULL,
            close_date TIMESTAMP NOT NULL,
            status TEXT DEFAULT 'draft' CHECK (status IN ('draft', 'open', 'closed')),
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subject_choice_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            subject_preferences TEXT,
            additional_notes TEXT,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (form_id) REFERENCES subject_choice_forms (id),
            FOREIGN KEY (staff_id) REFERENCES users (id),
            UNIQUE(form_id, staff_id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS department_queries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            priority TEXT DEFAULT 'medium' CHECK (priority IN ('low', 'medium', 'high')),
            status TEXT DEFAULT 'open' CHECK (status IN ('open', 'in_progress', 'resolved')),
            created_by INTEGER NOT NULL,
            assigned_to INTEGER,
            resolved_by INTEGER,
            resolution_notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id),
            FOREIGN KEY (assigned_to) REFERENCES users (id),
            FOREIGN KEY (resolved_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timetable_configurations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            period_duration INTEGER DEFAULT 60,
            periods_per_day INTEGER DEFAULT 7,
            college_start_time TEXT DEFAULT '09:00',
            college_end_time TEXT DEFAULT '17:00',
            break_times TEXT,
            working_days TEXT DEFAULT '["Monday","Tuesday","Wednesday","Thursday","Friday"]',
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS generated_timetables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            timetable_type TEXT NOT NULL CHECK (timetable_type IN ('student', 'staff', 'classroom', 'lab')),
            timetable_data TEXT,
            status TEXT DEFAULT 'draft' CHECK (status IN ('draft', 'approved', 'rejected')),
            generated_by INTEGER NOT NULL,
            approved_by INTEGER,
            generation_constraints TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (generated_by) REFERENCES users (id),
            FOREIGN KEY (approved_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS credentials_export (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            plain_password TEXT NOT NULL,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            exported BOOLEAN DEFAULT FALSE,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            recipient_type TEXT NOT NULL CHECK (recipient_type IN ('staff', 'dept_admin', 'all')),
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS staff_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_user_id INTEGER NOT NULL,
            to_user_id INTEGER NOT NULL,
            subject TEXT NOT NULL,
            message TEXT NOT NULL,
            message_type TEXT DEFAULT 'report' CHECK (message_type IN ('report', 'query', 'feedback')),
            status TEXT DEFAULT 'unread' CHECK (status IN ('unread', 'read', 'replied')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (from_user_id) REFERENCES users (id),
            FOREIGN KEY (to_user_id) REFERENCES users (id)
        )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS staff_registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        employee_id TEXT NOT NULL,
        college TEXT NOT NULL,
        faculty TEXT NOT NULL,
        campus TEXT NOT NULL,
        contact_number TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        username TEXT UNIQUE,
        password TEXT,
        staff_role TEXT CHECK (
            staff_role IN (
                'hod',
                'associate_professor',
                'assistant_professor',
                'professor',
                'professor_sg',
                'vp',
                'dean'
            )
        ),
        registered BOOLEAN DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
''')

    

    conn.commit()
    conn.close()

# Health Check
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'message': 'SRM Timetable AI Backend is running'}), 200

# AUTH ROUTES
@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')

        print(f"📥 Login attempt with email: {email}")

        if not email or not password:
            print("❌ Missing email or password")
            return jsonify({'success': False, 'error': 'Email and password are required'}), 400

        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT u.id, u.name, u.email, u.password_hash, u.role, u.department_id,
                   u.staff_role, u.subjects_selected, u.subjects_locked, d.name as department_name
            FROM users u
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE u.email = ?
        ''', (email,))
        user_data = cursor.fetchone()
        conn.close()

        if not user_data:
            print("❌ User not found in DB")
            return jsonify({'success': False, 'error': 'Invalid email or password'}), 401

        user_id, name, db_email, db_password_hash, role, dept_id, staff_role, selected_subjects, locked, dept_name = user_data

        print(f"✅ User record found: {db_email}")
        password_match = check_password_hash(db_password_hash, password)
        print(f"🔐 Password match: {password_match}")

        if not password_match:
            return jsonify({'success': False, 'error': 'Invalid email or password'}), 401

        user = {
            'id': str(user_id),
            'name': name,
            'email': db_email,
            'role': role,
            'department_id': str(dept_id) if dept_id else None,
            'staff_role': staff_role,
            'subjects_selected': selected_subjects.split(',') if selected_subjects else [],
            'subjects_locked': bool(locked),
            'department_name': dept_name
        }

        access_token = create_access_token(identity=str(user_id))
        print(f"🎟️ Token issued for {db_email}")

        return jsonify({'success': True, 'data': {'user': user, 'token': access_token}}), 200

    except Exception as e:
        print("🔥 Login Exception:", str(e))
        return jsonify({'success': False, 'error': 'Login failed'}), 500


@app.route('/api/auth/verify', methods=['GET'])
@jwt_required()
def verify_token():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT u.id, u.name, u.email, u.role, u.department_id,
                   u.staff_role, u.subjects_selected, u.subjects_locked, d.name as department_name
            FROM users u
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE u.id = ?
        ''', (current_user_id,))

        user_data = cursor.fetchone()
        conn.close()

        if not user_data:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        user = {
            'id': str(user_data[0]),
            'name': user_data[1],
            'email': user_data[2],
            'role': user_data[3],
            'department_id': str(user_data[4]) if user_data[4] else None,
            'staff_role': user_data[5],
            'subjects_selected': user_data[6].split(',') if user_data[6] else [],
            'subjects_locked': bool(user_data[7]),
            'department_name': user_data[8]
        }

        return jsonify({'success': True, 'data': {'user': user}}), 200

    except Exception as e:
        return jsonify({'success': False, 'error': 'Token verification failed'}), 401


@app.route('/api/auth/logout', methods=['POST'])
@jwt_required()
def logout():
    return jsonify({'success': True, 'message': 'Logged out successfully'}), 200

# Get all users
@app.route('/api/users', methods=['GET'])
@jwt_required()
def get_users():
    try:
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT u.id, u.name, u.email, u.role, d.name as department_name
            FROM users u
            LEFT JOIN departments d ON u.department_id = d.id
            ORDER BY u.name
        ''')

        users_data = cursor.fetchall()
        conn.close()

        users_list = []
        for user in users_data:
            users_list.append({
                'id': str(user[0]),
                'name': user[1],
                'email': user[2],
                'role': user[3],
                'department_name': user[4]
            })

        return jsonify(users_list), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== ENHANCED ADMIN ROUTES ====================

def generate_secure_credentials(user_id, name, email):
    """Generate secure username and password for a user"""
    username = email.split('@')[0].lower()
    password_chars = string.ascii_letters + string.digits + "!@#$%^&*"
    plain_password = ''.join(secrets.choice(password_chars) for _ in range(10))
    password_hash = generate_password_hash(plain_password)
    
    conn = sqlite3.connect('timetable.db')
    cursor = conn.cursor()
    
    cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (password_hash, user_id))
    cursor.execute('INSERT INTO credentials_export (user_id, username, plain_password) VALUES (?, ?, ?)', 
                  (user_id, username, plain_password))
    
    conn.commit()
    conn.close()
    
    return username, plain_password

@app.route('/api/admin/staff-requests', methods=['GET'])
@jwt_required()
def get_staff_requests():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role[0] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('''
            SELECT sr.*, d.name as department_name, u.name as requested_by_name
            FROM staff_registration_requests sr
            JOIN departments d ON sr.department_id = d.id
            JOIN users u ON sr.requested_by = u.id
            ORDER BY sr.created_at DESC
        ''')
        
        requests_data = cursor.fetchall()
        conn.close()
        
        requests_list = []
        for req in requests_data:
            requests_list.append({
                'id': str(req[0]),
                'name': req[1],
                'employee_id': req[2],
                'email': req[3],
                'department_id': str(req[4]),
                'staff_role': req[5],
                'contact_number': req[6],
                'requested_by': str(req[7]),
                'status': req[8],
                'approved_by': str(req[9]) if req[9] else None,
                'credentials_generated': bool(req[10]),
                'created_at': req[11],
                'approved_at': req[12],
                'department_name': req[13],
                'requested_by_name': req[14]
            })
        
        return jsonify({'success': True, 'requests': requests_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/staff-requests', methods=['POST'])
@jwt_required()
def create_staff_request():
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role, department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if not user_data or user_data[0] != 'dept_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('''
            INSERT INTO staff_registration_requests 
            (name, employee_id, email, department_id, staff_role, contact_number, requested_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'], data['employee_id'], data['email'],
            user_data[1], data['staff_role'],
            data.get('contact_number'), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Staff registration request submitted'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/staff-requests/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def approve_staff_request(request_id):
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role[0] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('SELECT * FROM staff_registration_requests WHERE id = ?', (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'error': 'Request not found'}), 404
        
        # Create user
        cursor.execute('''
            INSERT INTO users (name, email, role, department_id, staff_role)
            VALUES (?, ?, 'staff', ?, ?)
        ''', (request_data[1], request_data[3], request_data[4], request_data[5]))
        
        user_id = cursor.lastrowid
        
        # Generate credentials
        username, password = generate_secure_credentials(user_id, request_data[1], request_data[3])
        
        # Update request status
        cursor.execute('''
            UPDATE staff_registration_requests 
            SET status = 'approved', approved_by = ?, approved_at = CURRENT_TIMESTAMP, credentials_generated = TRUE
            WHERE id = ?
        ''', (current_user_id, request_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Staff approved and credentials generated',
            'credentials': {'username': username, 'password': password}
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/credentials/export', methods=['GET'])
@jwt_required()
def export_credentials():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role[0] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('''
            SELECT ce.username, ce.plain_password, u.name, u.email, u.role, 
                   d.name as department_name, ce.generated_at
            FROM credentials_export ce
            JOIN users u ON ce.user_id = u.id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE ce.exported = FALSE
            ORDER BY ce.generated_at DESC
        ''')
        
        credentials_data = cursor.fetchall()
        
        if not credentials_data:
            return jsonify({'error': 'No new credentials to export'}), 404
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Credentials"
        
        headers = ['Name', 'Email', 'Role', 'Department', 'Username', 'Password', 'Generated At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, data in enumerate(credentials_data, 2):
            ws.cell(row=row, column=1, value=data[2])
            ws.cell(row=row, column=2, value=data[3])
            ws.cell(row=row, column=3, value=data[4].replace('_', ' ').title())
            ws.cell(row=row, column=4, value=data[5] or 'N/A')
            ws.cell(row=row, column=5, value=data[0])
            ws.cell(row=row, column=6, value=data[1])
            ws.cell(row=row, column=7, value=data[6])
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        cursor.execute('UPDATE credentials_export SET exported = TRUE WHERE exported = FALSE')
        conn.commit()
        conn.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'user_credentials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/analytics', methods=['GET'])
@jwt_required()
def analytics_summary():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role[0] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        analytics = {}
        
        cursor.execute('SELECT COUNT(*) as count FROM departments')
        analytics['total_departments'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM users WHERE role = "staff"')
        analytics['total_staff'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM staff_registration_requests WHERE status = "pending"')
        analytics['pending_approvals'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM generated_timetables')
        analytics['timetable_generations'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM users WHERE role = "dept_admin"')
        analytics['total_dept_admins'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM credentials_export WHERE exported = FALSE')
        analytics['pending_credentials'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) as count FROM notifications')
        analytics['total_notifications'] = cursor.fetchone()[0]
        
        conn.close()
        
        return jsonify({'success': True, 'analytics': analytics}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/notifications/send', methods=['GET', 'POST'])
@jwt_required()
def handle_notifications():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role[0] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        if request.method == 'POST':
            data = request.get_json()
            title = data.get('title', '').strip()
            message = data.get('message', '').strip()
            recipient_type = data.get('recipient_type', '')
            
            if not title or not message or not recipient_type:
                return jsonify({'error': 'Title, message, and recipient type are required'}), 400
            
            cursor.execute('''
                INSERT INTO notifications (title, message, recipient_type, created_by)
                VALUES (?, ?, ?, ?)
            ''', (title, message, recipient_type, current_user_id))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Notification sent successfully'}), 200
        
        # GET request
        cursor.execute('''
            SELECT n.*, u.name as created_by_name
            FROM notifications n
            JOIN users u ON n.created_by = u.id
            ORDER BY n.created_at DESC
            LIMIT 10
        ''')
        
        recent_notifications = cursor.fetchall()
        conn.close()
        
        notifications_list = []
        for notif in recent_notifications:
            notifications_list.append({
                'id': str(notif[0]),
                'title': notif[1],
                'message': notif[2],
                'recipient_type': notif[3],
                'created_by': str(notif[4]),
                'created_at': notif[5],
                'created_by_name': notif[6]
            })
        
        return jsonify({'success': True, 'recent_notifications': notifications_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== ENHANCED DEPARTMENT ADMIN ROUTES ====================

@app.route('/api/enhanced-admin/constraints', methods=['GET'])
@jwt_required()
def get_enhanced_constraints():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id, role FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if user_data[1] == 'main_admin':
            cursor.execute('''
                SELECT ec.*, d.name as department_name
                FROM enhanced_constraints ec
                JOIN departments d ON ec.department_id = d.id
                ORDER BY ec.created_at DESC
            ''')
        else:
            cursor.execute('''
                SELECT ec.*, d.name as department_name
                FROM enhanced_constraints ec
                JOIN departments d ON ec.department_id = d.id
                WHERE ec.department_id = ?
                ORDER BY ec.created_at DESC
            ''', (user_data[0],))
        
        constraints = cursor.fetchall()
        conn.close()
        
        constraints_list = []
        for constraint in constraints:
            constraints_list.append({
                'id': str(constraint[0]),
                'department_id': str(constraint[1]),
                'role': constraint[2],
                'max_subjects': constraint[3],
                'max_hours_per_week': constraint[4],
                'subject_types': json.loads(constraint[5]) if constraint[5] else [],
                'lab_faculty_required': constraint[6],
                'created_by': str(constraint[7]),
                'created_at': constraint[8],
                'department_name': constraint[9]
            })
        
        return jsonify({'success': True, 'constraints': constraints_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/constraints', methods=['POST'])
@jwt_required()
def create_enhanced_constraint():
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id, role FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if user_data[1] not in ['dept_admin', 'main_admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        department_id = data.get('department_id') if user_data[1] == 'main_admin' else user_data[0]
        
        cursor.execute('''
            INSERT INTO enhanced_constraints 
            (department_id, role, max_subjects, max_hours_per_week, subject_types, lab_faculty_required, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            department_id, data['role'], data['max_subjects'],
            data['max_hours_per_week'], json.dumps(data.get('subject_types', [])),
            data.get('lab_faculty_required', 1), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Constraint created successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/choice-forms', methods=['GET'])
@jwt_required()
def get_choice_forms():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            SELECT scf.*, u.name as created_by_name,
                   COUNT(scs.id) as submission_count
            FROM subject_choice_forms scf
            JOIN users u ON scf.created_by = u.id
            LEFT JOIN subject_choice_submissions scs ON scf.id = scs.form_id
            WHERE scf.department_id = ?
            GROUP BY scf.id
            ORDER BY scf.created_at DESC
        ''', (user_data[0],))
        
        forms = cursor.fetchall()
        conn.close()
        
        forms_list = []
        for form in forms:
            forms_list.append({
                'id': str(form[0]),
                'department_id': str(form[1]),
                'title': form[2],
                'description': form[3],
                'open_date': form[4],
                'close_date': form[5],
                'status': form[6],
                'created_by': str(form[7]),
                'created_at': form[8],
                'created_by_name': form[9],
                'submission_count': form[10]
            })
        
        return jsonify({'success': True, 'forms': forms_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/choice-forms', methods=['POST'])
@jwt_required()
def create_choice_form():
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            INSERT INTO subject_choice_forms 
            (department_id, title, description, open_date, close_date, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            user_data[0], data['title'], data['description'],
            data['open_date'], data['close_date'], current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Choice form created successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/choice-forms/<int:form_id>/toggle', methods=['POST'])
@jwt_required()
def toggle_choice_form(form_id):
    try:
        data = request.get_json()
        status = data.get('status')
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('UPDATE subject_choice_forms SET status = ? WHERE id = ?', (status, form_id))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Form status updated to {status}'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/queries', methods=['GET'])
@jwt_required()
def get_department_queries():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if user_role[0] == 'main_admin':
            cursor.execute('''
                SELECT dq.*, d.name as department_name, u.name as created_by_name
                FROM department_queries dq
                JOIN departments d ON dq.department_id = d.id
                JOIN users u ON dq.created_by = u.id
                ORDER BY dq.created_at DESC
            ''')
        else:
            cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
            user_data = cursor.fetchone()
            
            cursor.execute('''
                SELECT dq.*, d.name as department_name, u.name as created_by_name
                FROM department_queries dq
                JOIN departments d ON dq.department_id = d.id
                JOIN users u ON dq.created_by = u.id
                WHERE dq.department_id = ?
                ORDER BY dq.created_at DESC
            ''', (user_data[0],))
        
        queries = cursor.fetchall()
        conn.close()
        
        queries_list = []
        for query in queries:
            queries_list.append({
                'id': str(query[0]),
                'department_id': str(query[1]),
                'title': query[2],
                'description': query[3],
                'priority': query[4],
                'status': query[5],
                'created_by': str(query[6]),
                'assigned_to': str(query[7]) if query[7] else None,
                'resolved_by': str(query[8]) if query[8] else None,
                'resolution_notes': query[9],
                'created_at': query[10],
                'resolved_at': query[11],
                'department_name': query[12],
                'created_by_name': query[13]
            })
        
        return jsonify({'success': True, 'queries': queries_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/queries', methods=['POST'])
@jwt_required()
def create_query():
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            INSERT INTO department_queries 
            (department_id, title, description, priority, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            user_data[0], data['title'], data['description'],
            data.get('priority', 'medium'), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Query submitted successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-admin/queries/<int:query_id>/resolve', methods=['POST'])
@jwt_required()
def resolve_query(query_id):
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE department_queries 
            SET status = 'resolved', resolved_by = ?, resolution_notes = ?, resolved_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (current_user_id, data.get('resolution_notes'), query_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Query resolved successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== ENHANCED STAFF ROUTES ====================

@app.route('/api/enhanced-staff/choice-forms/available', methods=['GET'])
@jwt_required()
def get_available_choice_forms():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({'error': 'User not found'}), 404
        
        cursor.execute('''
            SELECT scf.*, 
                   CASE WHEN scs.id IS NOT NULL THEN 1 ELSE 0 END as has_submitted
            FROM subject_choice_forms scf
            LEFT JOIN subject_choice_submissions scs ON scf.id = scs.form_id AND scs.staff_id = ?
            WHERE scf.department_id = ? AND scf.status = 'open'
            ORDER BY scf.close_date ASC
        ''', (current_user_id, user_data[0]))
        
        forms = cursor.fetchall()
        conn.close()
        
        forms_list = []
        for form in forms:
            forms_list.append({
                'id': str(form[0]),
                'department_id': str(form[1]),
                'title': form[2],
                'description': form[3],
                'open_date': form[4],
                'close_date': form[5],
                'status': form[6],
                'created_by': str(form[7]),
                'created_at': form[8],
                'has_submitted': bool(form[9])
            })
        
        return jsonify({'success': True, 'forms': forms_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-staff/choice-forms/<int:form_id>/submit', methods=['POST'])
@jwt_required()
def submit_choice_form(form_id):
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('SELECT status FROM subject_choice_forms WHERE id = ?', (form_id,))
        form_data = cursor.fetchone()
        
        if not form_data or form_data[0] != 'open':
            return jsonify({'error': 'Form is not available for submission'}), 400
        
        cursor.execute('''
            INSERT OR REPLACE INTO subject_choice_submissions 
            (form_id, staff_id, subject_preferences, additional_notes)
            VALUES (?, ?, ?, ?)
        ''', (
            form_id, current_user_id, 
            json.dumps(data.get('subject_preferences', [])),
            data.get('additional_notes', '')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Preferences submitted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-staff/my-submissions', methods=['GET'])
@jwt_required()
def get_my_submissions():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT scs.*, scf.title as form_title, scf.status as form_status
            FROM subject_choice_submissions scs
            JOIN subject_choice_forms scf ON scs.form_id = scf.id
            WHERE scs.staff_id = ?
            ORDER BY scs.submitted_at DESC
        ''', (current_user_id,))
        
        submissions = cursor.fetchall()
        conn.close()
        
        submissions_list = []
        for submission in submissions:
            submissions_list.append({
                'id': str(submission[0]),
                'form_id': str(submission[1]),
                'staff_id': str(submission[2]),
                'subject_preferences': submission[3],
                'additional_notes': submission[4],
                'submitted_at': submission[5],
                'form_title': submission[6],
                'form_status': submission[7]
            })
        
        return jsonify({'success': True, 'submissions': submissions_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enhanced-staff/my-timetable', methods=['GET'])
@jwt_required()
def get_my_timetable():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT gt.timetable_data
            FROM generated_timetables gt
            JOIN users u ON gt.department_id = u.department_id
            WHERE u.id = ? AND gt.timetable_type = 'staff' AND gt.status = 'approved'
            ORDER BY gt.created_at DESC
            LIMIT 1
        ''', (current_user_id,))
        
        timetable_data = cursor.fetchone()
        
        if timetable_data:
            timetable_json = json.loads(timetable_data[0])
            staff_timetable = timetable_json.get(str(current_user_id), {})
        else:
            staff_timetable = {}
        
        conn.close()
        
        return jsonify({'success': True, 'timetable': staff_timetable}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/staff/messages', methods=['GET', 'POST'])
@jwt_required()
def handle_staff_messages():
    try:
        current_user_id = get_jwt_identity()
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        if request.method == 'POST':
            data = request.get_json()
            
            cursor.execute('''
                INSERT INTO staff_messages 
                (from_user_id, to_user_id, subject, message, message_type)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                current_user_id, data['to_user_id'], data['subject'],
                data['message'], data.get('message_type', 'report')
            ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Message sent successfully'}), 201
        
        # GET request - get messages for current user
        cursor.execute('''
            SELECT sm.*, u1.name as from_name, u2.name as to_name
            FROM staff_messages sm
            JOIN users u1 ON sm.from_user_id = u1.id
            JOIN users u2 ON sm.to_user_id = u2.id
            WHERE sm.from_user_id = ? OR sm.to_user_id = ?
            ORDER BY sm.created_at DESC
        ''', (current_user_id, current_user_id))
        
        messages = cursor.fetchall()
        conn.close()
        
        messages_list = []
        for msg in messages:
            messages_list.append({
                'id': str(msg[0]),
                'from_user_id': str(msg[1]),
                'to_user_id': str(msg[2]),
                'subject': msg[3],
                'message': msg[4],
                'message_type': msg[5],
                'status': msg[6],
                'created_at': msg[7],
                'from_name': msg[8],
                'to_name': msg[9]
            })
        
        return jsonify({'success': True, 'messages': messages_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Timetable stats
@app.route('/api/timetables/stats', methods=['GET'])
@jwt_required()
def get_timetable_stats():
    try:
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()

        cursor.execute('SELECT COUNT(*) FROM timetables')
        total = cursor.fetchone()[0]
        conn.close()

        return jsonify({'total': total}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/admin/staff-registrations', methods=['GET'])
def list_registered_staff():
    conn = sqlite3.connect('timetable.db')
    cursor = conn.cursor()
    cursor.execute('SELECT name, email, staff_role, registered FROM staff_registrations ORDER BY created_at DESC')
    rows = cursor.fetchall()
    conn.close()

    result = [
        {
            'name': r[0],
            'email': r[1],
            'staff_role': r[2],
            'registered': bool(r[3])
        } for r in rows
    ]
    return jsonify(result), 200


@app.route('/api/admin/register-staff', methods=['POST'])
def register_staff():
    try:
        data = request.get_json()
        name = data.get('name')
        employee_id = data.get('employee_id')
        college = data.get('college')
        faculty = data.get('faculty')
        campus = data.get('campus')
        contact_number = data.get('contact_number')
        email = data.get('email')
        staff_role = data.get('staff_role')

        if not all([name, employee_id, college, faculty, campus, contact_number, email, staff_role]):
            return jsonify({'success': False, 'error': 'All fields are required'}), 400

        # Validate role
        valid_roles = [
            'hod',
            'associate_professor',
            'assistant_professor',
            'professor',
            'professor_sg',
            'vp',
            'dean'
        ]
        if staff_role not in valid_roles:
            return jsonify({'success': False, 'error': f'Invalid staff role: {staff_role}'}), 400

        # Generate credentials
        username = email.split('@')[0]
        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO staff_registrations (
                name, employee_id, college, faculty, campus,
                contact_number, email, username, password, staff_role, registered
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            name, employee_id, college, faculty, campus,
            contact_number, email, username, password, staff_role, 0
        ))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'Staff registered successfully',
            'credentials': {
                'username': username,
                'password': password
            }
        }), 201

    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'User already registered'}), 409
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
# Run App
if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
