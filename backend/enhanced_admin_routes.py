from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash
import sqlite3
import secrets
import string
import hashlib
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import os
import json
import requests

enhanced_admin_bp = Blueprint('enhanced_admin', __name__, url_prefix='/api/enhanced-admin')

def get_db_connection():
    """Get database connection"""
    conn = sqlite3.connect('timetable.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_enhanced_tables():
    """Initialize enhanced tables"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Staff registration requests table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS staff_registration_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            employee_id TEXT NOT NULL,
            email TEXT NOT NULL,
            department_id INTEGER NOT NULL,
            staff_role TEXT NOT NULL,
            contact_number TEXT,
            requested_by INTEGER NOT NULL,
            status TEXT DEFAULT 'pending' CHECK (status IN ('pending', 'approved', 'rejected')),
            approved_by INTEGER,
            credentials_generated BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (requested_by) REFERENCES users (id),
            FOREIGN KEY (approved_by) REFERENCES users (id)
        )
    ''')
    
    # Enhanced constraints table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS enhanced_constraints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            max_subjects INTEGER NOT NULL DEFAULT 1,
            max_hours_per_week INTEGER NOT NULL DEFAULT 8,
            subject_types TEXT, -- JSON array of allowed subject types
            lab_faculty_required INTEGER DEFAULT 1,
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Subject choice forms table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subject_choice_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            open_date TIMESTAMP NOT NULL,
            close_date TIMESTAMP NOT NULL,
            status TEXT DEFAULT 'draft' CHECK (status IN ('draft', 'open', 'closed')),
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Subject choices submissions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subject_choice_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            subject_preferences TEXT, -- JSON array of subject IDs in preference order
            additional_notes TEXT,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (form_id) REFERENCES subject_choice_forms (id),
            FOREIGN KEY (staff_id) REFERENCES users (id),
            UNIQUE(form_id, staff_id)
        )
    ''')
    
    # Department queries table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS department_queries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            priority TEXT DEFAULT 'medium' CHECK (priority IN ('low', 'medium', 'high')),
            status TEXT DEFAULT 'open' CHECK (status IN ('open', 'in_progress', 'resolved')),
            created_by INTEGER NOT NULL,
            assigned_to INTEGER,
            resolved_by INTEGER,
            resolution_notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id),
            FOREIGN KEY (assigned_to) REFERENCES users (id),
            FOREIGN KEY (resolved_by) REFERENCES users (id)
        )
    ''')
    
    # Timetable configurations table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timetable_configurations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            period_duration INTEGER DEFAULT 60, -- minutes
            periods_per_day INTEGER DEFAULT 7,
            college_start_time TEXT DEFAULT '09:00',
            college_end_time TEXT DEFAULT '17:00',
            break_times TEXT, -- JSON array of break times
            working_days TEXT DEFAULT '["Monday","Tuesday","Wednesday","Thursday","Friday"]',
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Generated timetables table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS generated_timetables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            timetable_type TEXT NOT NULL CHECK (timetable_type IN ('student', 'staff', 'classroom', 'lab')),
            timetable_data TEXT, -- JSON data
            status TEXT DEFAULT 'draft' CHECK (status IN ('draft', 'approved', 'rejected')),
            generated_by INTEGER NOT NULL,
            approved_by INTEGER,
            generation_constraints TEXT, -- JSON of constraints used
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (generated_by) REFERENCES users (id),
            FOREIGN KEY (approved_by) REFERENCES users (id)
        )
    ''')
    
    # Credentials export table (from previous enhancement)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS credentials_export (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            plain_password TEXT NOT NULL,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            exported BOOLEAN DEFAULT FALSE,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    conn.commit()
    conn.close()

# Initialize tables when module is imported
init_enhanced_tables()

# Staff Registration Routes
@enhanced_admin_bp.route('/staff-requests', methods=['GET'])
@jwt_required()
def get_staff_requests():
    """Get staff registration requests"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if main admin
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('''
            SELECT sr.*, d.name as department_name, u.name as requested_by_name
            FROM staff_registration_requests sr
            JOIN departments d ON sr.department_id = d.id
            JOIN users u ON sr.requested_by = u.id
            ORDER BY sr.created_at DESC
        ''')
        
        requests_data = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'requests': [dict(row) for row in requests_data]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/staff-requests', methods=['POST'])
@jwt_required()
def create_staff_request():
    """Create staff registration request (Department Admin)"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verify department admin
        cursor.execute('SELECT role, department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if not user_data or user_data['role'] != 'dept_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        cursor.execute('''
            INSERT INTO staff_registration_requests 
            (name, employee_id, email, department_id, staff_role, contact_number, requested_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'], data['employee_id'], data['email'],
            user_data['department_id'], data['staff_role'],
            data.get('contact_number'), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Staff registration request submitted'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/staff-requests/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def approve_staff_request(request_id):
    """Approve staff registration request and generate credentials"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verify main admin
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get request details
        cursor.execute('SELECT * FROM staff_registration_requests WHERE id = ?', (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'error': 'Request not found'}), 404
        
        # Generate credentials
        username = request_data['email'].split('@')[0].lower()
        password = ''.join(secrets.choice(string.ascii_letters + string.digits + "!@#$%^&*") for _ in range(10))
        password_hash = generate_password_hash(password)
        
        # Create user
        cursor.execute('''
            INSERT INTO users (name, email, password_hash, role, department_id, staff_role)
            VALUES (?, ?, ?, 'staff', ?, ?)
        ''', (
            request_data['name'], request_data['email'], password_hash,
            request_data['department_id'], request_data['staff_role']
        ))
        
        user_id = cursor.lastrowid
        
        # Store credentials for export
        cursor.execute('''
            INSERT INTO credentials_export (user_id, username, plain_password)
            VALUES (?, ?, ?)
        ''', (user_id, username, password))
        
        # Update request status
        cursor.execute('''
            UPDATE staff_registration_requests 
            SET status = 'approved', approved_by = ?, approved_at = CURRENT_TIMESTAMP, credentials_generated = TRUE
            WHERE id = ?
        ''', (current_user_id, request_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Staff approved and credentials generated',
            'credentials': {'username': username, 'password': password}
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Enhanced Constraints Routes
@enhanced_admin_bp.route('/constraints', methods=['GET'])
@jwt_required()
def get_enhanced_constraints():
    """Get enhanced constraints for department"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id, role FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if user_data['role'] == 'main_admin':
            cursor.execute('''
                SELECT ec.*, d.name as department_name
                FROM enhanced_constraints ec
                JOIN departments d ON ec.department_id = d.id
                ORDER BY ec.created_at DESC
            ''')
        else:
            cursor.execute('''
                SELECT ec.*, d.name as department_name
                FROM enhanced_constraints ec
                JOIN departments d ON ec.department_id = d.id
                WHERE ec.department_id = ?
                ORDER BY ec.created_at DESC
            ''', (user_data['department_id'],))
        
        constraints = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'constraints': [dict(row) for row in constraints]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/constraints', methods=['POST'])
@jwt_required()
def create_enhanced_constraint():
    """Create enhanced constraint"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id, role FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        if user_data['role'] not in ['dept_admin', 'main_admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        department_id = data.get('department_id') if user_data['role'] == 'main_admin' else user_data['department_id']
        
        cursor.execute('''
            INSERT INTO enhanced_constraints 
            (department_id, role, max_subjects, max_hours_per_week, subject_types, lab_faculty_required, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            department_id, data['role'], data['max_subjects'],
            data['max_hours_per_week'], json.dumps(data.get('subject_types', [])),
            data.get('lab_faculty_required', 1), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Constraint created successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Subject Choice Forms Routes
@enhanced_admin_bp.route('/choice-forms', methods=['GET'])
@jwt_required()
def get_choice_forms():
    """Get subject choice forms"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            SELECT scf.*, u.name as created_by_name,
                   COUNT(scs.id) as submission_count
            FROM subject_choice_forms scf
            JOIN users u ON scf.created_by = u.id
            LEFT JOIN subject_choice_submissions scs ON scf.id = scs.form_id
            WHERE scf.department_id = ?
            GROUP BY scf.id
            ORDER BY scf.created_at DESC
        ''', (user_data['department_id'],))
        
        forms = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'forms': [dict(row) for row in forms]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/choice-forms', methods=['POST'])
@jwt_required()
def create_choice_form():
    """Create subject choice form"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            INSERT INTO subject_choice_forms 
            (department_id, title, description, open_date, close_date, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            user_data['department_id'], data['title'], data['description'],
            data['open_date'], data['close_date'], current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Choice form created successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/choice-forms/<int:form_id>/toggle', methods=['POST'])
@jwt_required()
def toggle_choice_form(form_id):
    """Toggle choice form status"""
    try:
        data = request.get_json()
        status = data.get('status')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('UPDATE subject_choice_forms SET status = ? WHERE id = ?', (status, form_id))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Form status updated to {status}'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Department Queries Routes
@enhanced_admin_bp.route('/queries', methods=['GET'])
@jwt_required()
def get_department_queries():
    """Get department queries"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if user_role['role'] == 'main_admin':
            cursor.execute('''
                SELECT dq.*, d.name as department_name, u.name as created_by_name
                FROM department_queries dq
                JOIN departments d ON dq.department_id = d.id
                JOIN users u ON dq.created_by = u.id
                ORDER BY dq.created_at DESC
            ''')
        else:
            cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
            user_data = cursor.fetchone()
            
            cursor.execute('''
                SELECT dq.*, d.name as department_name, u.name as created_by_name
                FROM department_queries dq
                JOIN departments d ON dq.department_id = d.id
                JOIN users u ON dq.created_by = u.id
                WHERE dq.department_id = ?
                ORDER BY dq.created_at DESC
            ''', (user_data['department_id'],))
        
        queries = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'queries': [dict(row) for row in queries]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/queries', methods=['POST'])
@jwt_required()
def create_query():
    """Create department query"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            INSERT INTO department_queries 
            (department_id, title, description, priority, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            user_data['department_id'], data['title'], data['description'],
            data.get('priority', 'medium'), current_user_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Query submitted successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/queries/<int:query_id>/resolve', methods=['POST'])
@jwt_required()
def resolve_query(query_id):
    """Resolve department query"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE department_queries 
            SET status = 'resolved', resolved_by = ?, resolution_notes = ?, resolved_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (current_user_id, data.get('resolution_notes'), query_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Query resolved successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# AI Timetable Generation Routes
@enhanced_admin_bp.route('/timetable/generate', methods=['POST'])
@jwt_required()
def generate_ai_timetable():
    """Generate AI-powered timetable"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        department_id = user_data['department_id']
        
        # Get all constraints and data
        cursor.execute('SELECT * FROM enhanced_constraints WHERE department_id = ?', (department_id,))
        constraints = cursor.fetchall()
        
        cursor.execute('SELECT * FROM timetable_configurations WHERE department_id = ?', (department_id,))
        config = cursor.fetchone()
        
        cursor.execute('''
            SELECT u.*, GROUP_CONCAT(scs.subject_preferences) as preferences
            FROM users u
            LEFT JOIN subject_choice_submissions scs ON u.id = scs.staff_id
            WHERE u.department_id = ? AND u.role = 'staff'
            GROUP BY u.id
        ''', (department_id,))
        staff_data = cursor.fetchall()
        
        cursor.execute('SELECT * FROM subjects WHERE department_id = ?', (department_id,))
        subjects = cursor.fetchall()
        
        cursor.execute('SELECT * FROM classrooms WHERE department_id = ?', (department_id,))
        classrooms = cursor.fetchall()
        
        # Generate timetables using AI logic
        timetable_generator = AITimetableGenerator()
        generated_timetables = timetable_generator.generate_comprehensive_timetables(
            constraints, config, staff_data, subjects, classrooms
        )
        
        # Store generated timetables
        for timetable_type, timetable_data in generated_timetables.items():
            cursor.execute('''
                INSERT INTO generated_timetables 
                (department_id, timetable_type, timetable_data, generated_by, generation_constraints)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                department_id, timetable_type, json.dumps(timetable_data),
                current_user_id, json.dumps([dict(c) for c in constraints])
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Timetables generated successfully',
            'timetables': generated_timetables
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@enhanced_admin_bp.route('/timetables', methods=['GET'])
@jwt_required()
def get_generated_timetables():
    """Get generated timetables"""
    try:
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT department_id FROM users WHERE id = ?', (current_user_id,))
        user_data = cursor.fetchone()
        
        cursor.execute('''
            SELECT gt.*, u.name as generated_by_name
            FROM generated_timetables gt
            JOIN users u ON gt.generated_by = u.id
            WHERE gt.department_id = ?
            ORDER BY gt.created_at DESC
        ''', (user_data['department_id'],))
        
        timetables = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'timetables': [dict(row) for row in timetables]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# AI Timetable Generator Class
class AITimetableGenerator:
    def __init__(self):
        self.groq_api_key = os.getenv('GROQ_API_KEY')
        self.google_api_key = os.getenv('GOOGLE_API_KEY')
    
    def generate_comprehensive_timetables(self, constraints, config, staff_data, subjects, classrooms):
        """Generate all 4 types of timetables using AI"""
        
        # Prepare data for AI processing
        constraint_rules = self._process_constraints(constraints)
        staff_preferences = self._process_staff_preferences(staff_data)
        subject_requirements = self._process_subjects(subjects)
        classroom_availability = self._process_classrooms(classrooms)
        
        # Generate base timetable using constraint satisfaction
        base_timetable = self._generate_base_timetable(
            constraint_rules, staff_preferences, subject_requirements, classroom_availability, config
        )
        
        # Generate 4 different views
        timetables = {
            'student': self._generate_student_timetable(base_timetable),
            'staff': self._generate_staff_timetable(base_timetable),
            'classroom': self._generate_classroom_timetable(base_timetable),
            'lab': self._generate_lab_timetable(base_timetable)
        }
        
        return timetables
    
    def _process_constraints(self, constraints):
        """Process constraints into usable rules"""
        rules = {}
        for constraint in constraints:
            rules[constraint['role']] = {
                'max_subjects': constraint['max_subjects'],
                'max_hours': constraint['max_hours_per_week'],
                'subject_types': json.loads(constraint['subject_types']) if constraint['subject_types'] else [],
                'lab_faculty_required': constraint['lab_faculty_required']
            }
        return rules
    
    def _process_staff_preferences(self, staff_data):
        """Process staff preferences"""
        preferences = {}
        for staff in staff_data:
            if staff['preferences']:
                try:
                    prefs = json.loads(staff['preferences'])
                    preferences[staff['id']] = {
                        'name': staff['name'],
                        'role': staff['staff_role'],
                        'preferences': prefs
                    }
                except:
                    preferences[staff['id']] = {
                        'name': staff['name'],
                        'role': staff['staff_role'],
                        'preferences': []
                    }
        return preferences
    
    def _process_subjects(self, subjects):
        """Process subject requirements"""
        subject_reqs = {}
        for subject in subjects:
            subject_reqs[subject['id']] = {
                'name': subject['name'],
                'code': subject['code'],
                'credits': subject['credits'],
                'hours_per_week': subject['credits'] * 2  # Assume 2 hours per credit
            }
        return subject_reqs
    
    def _process_classrooms(self, classrooms):
        """Process classroom availability"""
        classroom_data = {}
        for classroom in classrooms:
            classroom_data[classroom['id']] = {
                'name': classroom['name'],
                'capacity': classroom['capacity'],
                'type': 'lab' if 'lab' in classroom['name'].lower() else 'classroom'
            }
        return classroom_data
    
    def _generate_base_timetable(self, constraints, staff_prefs, subjects, classrooms, config):
        """Generate base timetable using constraint satisfaction"""
        
        # Time slots based on configuration
        if config:
            periods_per_day = config['periods_per_day']
            working_days = json.loads(config['working_days'])
        else:
            periods_per_day = 7
            working_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        time_slots = [f"Period {i+1}" for i in range(periods_per_day)]
        
        # Initialize timetable structure
        timetable = []
        
        # Track assignments
        staff_workload = {staff_id: 0 for staff_id in staff_prefs.keys()}
        classroom_schedule = {day: {slot: None for slot in time_slots} for day in working_days}
        
        # Assign subjects based on preferences and constraints
        for staff_id, staff_info in staff_prefs.items():
            role = staff_info['role']
            max_hours = constraints.get(role, {}).get('max_hours', 8)
            
            for subject_id in staff_info['preferences']:
                if staff_workload[staff_id] >= max_hours:
                    break
                
                subject_info = subjects.get(int(subject_id), {})
                hours_needed = subject_info.get('hours_per_week', 3)
                
                # Find available slots
                assigned_hours = 0
                for day in working_days:
                    if assigned_hours >= hours_needed:
                        break
                    
                    for slot in time_slots:
                        if assigned_hours >= hours_needed:
                            break
                        
                        if classroom_schedule[day][slot] is None:
                            # Find available classroom
                            available_classroom = None
                            for classroom_id, classroom_info in classrooms.items():
                                if self._is_classroom_available(classroom_id, day, slot, timetable):
                                    available_classroom = classroom_id
                                    break
                            
                            if available_classroom:
                                entry = {
                                    'day': day,
                                    'time_slot': slot,
                                    'subject_id': subject_id,
                                    'staff_id': staff_id,
                                    'classroom_id': available_classroom,
                                    'subject_name': subject_info.get('name', ''),
                                    'staff_name': staff_info['name'],
                                    'classroom_name': classrooms[available_classroom]['name']
                                }
                                
                                timetable.append(entry)
                                classroom_schedule[day][slot] = entry
                                assigned_hours += 1
                                staff_workload[staff_id] += 1
        
        return timetable
    
    def _is_classroom_available(self, classroom_id, day, slot, timetable):
        """Check if classroom is available"""
        for entry in timetable:
            if (entry['classroom_id'] == classroom_id and 
                entry['day'] == day and 
                entry['time_slot'] == slot):
                return False
        return True
    
    def _generate_student_timetable(self, base_timetable):
        """Generate student view timetable"""
        student_timetable = {}
        
        for entry in base_timetable:
            day = entry['day']
            slot = entry['time_slot']
            
            if day not in student_timetable:
                student_timetable[day] = {}
            
            student_timetable[day][slot] = {
                'subject': entry['subject_name'],
                'staff': entry['staff_name'],
                'classroom': entry['classroom_name']
            }
        
        return student_timetable
    
    def _generate_staff_timetable(self, base_timetable):
        """Generate staff view timetable"""
        staff_timetable = {}
        
        for entry in base_timetable:
            staff_id = entry['staff_id']
            
            if staff_id not in staff_timetable:
                staff_timetable[staff_id] = {
                    'name': entry['staff_name'],
                    'schedule': {}
                }
            
            day = entry['day']
            if day not in staff_timetable[staff_id]['schedule']:
                staff_timetable[staff_id]['schedule'][day] = {}
            
            staff_timetable[staff_id]['schedule'][day][entry['time_slot']] = {
                'subject': entry['subject_name'],
                'classroom': entry['classroom_name']
            }
        
        return staff_timetable
    
    def _generate_classroom_timetable(self, base_timetable):
        """Generate classroom view timetable"""
        classroom_timetable = {}
        
        for entry in base_timetable:
            classroom_id = entry['classroom_id']
            
            if classroom_id not in classroom_timetable:
                classroom_timetable[classroom_id] = {
                    'name': entry['classroom_name'],
                    'schedule': {}
                }
            
            day = entry['day']
            if day not in classroom_timetable[classroom_id]['schedule']:
                classroom_timetable[classroom_id]['schedule'][day] = {}
            
            classroom_timetable[classroom_id]['schedule'][day][entry['time_slot']] = {
                'subject': entry['subject_name'],
                'staff': entry['staff_name']
            }
        
        return classroom_timetable
    
    def _generate_lab_timetable(self, base_timetable):
        """Generate lab view timetable"""
        lab_timetable = {}
        
        for entry in base_timetable:
            if 'lab' in entry['classroom_name'].lower():
                classroom_id = entry['classroom_id']
                
                if classroom_id not in lab_timetable:
                    lab_timetable[classroom_id] = {
                        'name': entry['classroom_name'],
                        'schedule': {}
                    }
                
                day = entry['day']
                if day not in lab_timetable[classroom_id]['schedule']:
                    lab_timetable[classroom_id]['schedule'][day] = {}
                
                lab_timetable[classroom_id]['schedule'][day][entry['time_slot']] = {
                    'subject': entry['subject_name'],
                    'staff': entry['staff_name']
                }
        
        return lab_timetable

# Export Routes
@enhanced_admin_bp.route('/export/choice-submissions/<int:form_id>', methods=['GET'])
@jwt_required()
def export_choice_submissions(form_id):
    """Export choice submissions to Excel"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT scs.*, u.name as staff_name, scf.title as form_title
            FROM subject_choice_submissions scs
            JOIN users u ON scs.staff_id = u.id
            JOIN subject_choice_forms scf ON scs.form_id = scf.id
            WHERE scs.form_id = ?
        ''', (form_id,))
        
        submissions = cursor.fetchall()
        conn.close()
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Choice Submissions"
        
        # Headers
        headers = ['Staff Name', 'Subject Preferences', 'Additional Notes', 'Submitted At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, submission in enumerate(submissions, 2):
            ws.cell(row=row, column=1, value=submission['staff_name'])
            ws.cell(row=row, column=2, value=submission['subject_preferences'])
            ws.cell(row=row, column=3, value=submission['additional_notes'] or '')
            ws.cell(row=row, column=4, value=submission['submitted_at'])
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'choice_submissions_{form_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500