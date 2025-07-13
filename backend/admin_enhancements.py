import os
import hashlib
import secrets
import string
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile

admin_bp = Blueprint('admin_enhancements', __name__, url_prefix='/admin')

def get_db_connection():
    """Get database connection"""
    conn = sqlite3.connect('timetable.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_enhancement_tables():
    """Initialize new tables for enhancements"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Credentials export table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS credentials_export (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            plain_password TEXT NOT NULL,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            exported BOOLEAN DEFAULT FALSE,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Notifications table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            recipient_type TEXT NOT NULL CHECK (recipient_type IN ('staff', 'dept_admin', 'all')),
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Syllabus uploads table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS syllabus_uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            department_id INTEGER,
            uploaded_by INTEGER NOT NULL,
            status TEXT DEFAULT 'pending' CHECK (status IN ('pending', 'approved', 'rejected')),
            reviewed_by INTEGER,
            review_notes TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reviewed_at TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (uploaded_by) REFERENCES users (id),
            FOREIGN KEY (reviewed_by) REFERENCES users (id)
        )
    ''')
    
    # Timetable generation logs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timetable_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            generation_type TEXT NOT NULL,
            generated_by INTEGER NOT NULL,
            status TEXT DEFAULT 'completed' CHECK (status IN ('completed', 'failed', 'in_progress')),
            entries_count INTEGER DEFAULT 0,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments (id),
            FOREIGN KEY (generated_by) REFERENCES users (id)
        )
    ''')
    
    conn.commit()
    conn.close()

# Initialize tables when module is imported
init_enhancement_tables()

def generate_secure_credentials(user_id, name, email):
    """Generate secure username and password for a user"""
    # Generate username from email
    username = email.split('@')[0].lower()
    
    # Generate random 10-character password
    password_chars = string.ascii_letters + string.digits + "!@#$%^&*"
    plain_password = ''.join(secrets.choice(password_chars) for _ in range(10))
    
    # Hash password using SHA256
    password_hash = hashlib.sha256(plain_password.encode()).hexdigest()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Update user table with hashed password
    cursor.execute('''
        UPDATE users SET password_hash = ? WHERE id = ?
    ''', (password_hash, user_id))
    
    # Store plain credentials in export table
    cursor.execute('''
        INSERT INTO credentials_export (user_id, username, plain_password)
        VALUES (?, ?, ?)
    ''', (user_id, username, plain_password))
    
    conn.commit()
    conn.close()
    
    return username, plain_password

@admin_bp.route('/credentials/generate', methods=['POST'])
@jwt_required()
def generate_credentials():
    """Generate credentials for approved staff and department admins"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get users without credentials (staff and dept_admin)
        cursor.execute('''
            SELECT u.id, u.name, u.email, u.role 
            FROM users u 
            LEFT JOIN credentials_export ce ON u.id = ce.user_id 
            WHERE u.role IN ('staff', 'dept_admin') 
            AND ce.user_id IS NULL 
            AND u.password_hash IS NULL
        ''')
        
        users_without_credentials = cursor.fetchall()
        conn.close()
        
        generated_count = 0
        for user in users_without_credentials:
            username, password = generate_secure_credentials(
                user['id'], user['name'], user['email']
            )
            generated_count += 1
        
        return jsonify({
            'success': True,
            'message': f'Generated credentials for {generated_count} users',
            'count': generated_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/credentials/export')
@jwt_required()
def export_credentials():
    """Export generated credentials as Excel file"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get credentials data
        cursor.execute('''
            SELECT ce.username, ce.plain_password, u.name, u.email, u.role, 
                   d.name as department_name, ce.generated_at
            FROM credentials_export ce
            JOIN users u ON ce.user_id = u.id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE ce.exported = FALSE
            ORDER BY ce.generated_at DESC
        ''')
        
        credentials_data = cursor.fetchall()
        
        if not credentials_data:
            return jsonify({'error': 'No new credentials to export'}), 404
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Credentials"
        
        # Headers
        headers = ['Name', 'Email', 'Role', 'Department', 'Username', 'Password', 'Generated At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, data in enumerate(credentials_data, 2):
            ws.cell(row=row, column=1, value=data['name'])
            ws.cell(row=row, column=2, value=data['email'])
            ws.cell(row=row, column=3, value=data['role'].replace('_', ' ').title())
            ws.cell(row=row, column=4, value=data['department_name'] or 'N/A')
            ws.cell(row=row, column=5, value=data['username'])
            ws.cell(row=row, column=6, value=data['plain_password'])
            ws.cell(row=row, column=7, value=data['generated_at'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        # Mark as exported
        cursor.execute('UPDATE credentials_export SET exported = TRUE WHERE exported = FALSE')
        conn.commit()
        conn.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'user_credentials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/analytics')
@jwt_required()
def analytics_summary():
    """Get analytics summary for main admin dashboard"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get analytics data
        analytics = {}
        
        # Total departments
        cursor.execute('SELECT COUNT(*) as count FROM departments')
        analytics['total_departments'] = cursor.fetchone()['count']
        
        # Total staff
        cursor.execute('SELECT COUNT(*) as count FROM users WHERE role = "staff"')
        analytics['total_staff'] = cursor.fetchone()['count']
        
        # Pending approvals (syllabus uploads)
        cursor.execute('SELECT COUNT(*) as count FROM syllabus_uploads WHERE status = "pending"')
        analytics['pending_approvals'] = cursor.fetchone()['count']
        
        # Timetable generations
        cursor.execute('SELECT COUNT(*) as count FROM timetable_logs')
        analytics['timetable_generations'] = cursor.fetchone()['count']
        
        # Additional metrics
        cursor.execute('SELECT COUNT(*) as count FROM users WHERE role = "dept_admin"')
        analytics['total_dept_admins'] = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM credentials_export WHERE exported = FALSE')
        analytics['pending_credentials'] = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM notifications')
        analytics['total_notifications'] = cursor.fetchone()['count']
        
        conn.close()
        
        return jsonify({
            'success': True,
            'analytics': analytics
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/notifications/send', methods=['GET', 'POST'])
@jwt_required()
def send_notification():
    """Send notifications to users"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        if request.method == 'POST':
            data = request.get_json()
            title = data.get('title', '').strip()
            message = data.get('message', '').strip()
            recipient_type = data.get('recipient_type', '')
            
            if not title or not message or not recipient_type:
                return jsonify({'error': 'Title, message, and recipient type are required'}), 400
            
            if recipient_type not in ['staff', 'dept_admin', 'all']:
                return jsonify({'error': 'Invalid recipient type'}), 400
            
            # Insert notification
            cursor.execute('''
                INSERT INTO notifications (title, message, recipient_type, created_by)
                VALUES (?, ?, ?, ?)
            ''', (title, message, recipient_type, current_user_id))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'Notification sent successfully'
            })
        
        # GET request - return notification form data
        cursor.execute('''
            SELECT n.*, u.name as created_by_name
            FROM notifications n
            JOIN users u ON n.created_by = u.id
            ORDER BY n.created_at DESC
            LIMIT 10
        ''')
        
        recent_notifications = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'recent_notifications': [dict(row) for row in recent_notifications]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/syllabus/review')
@jwt_required()
def syllabus_review():
    """Review uploaded syllabus files"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get syllabus uploads
        cursor.execute('''
            SELECT su.*, u.name as uploaded_by_name, d.name as department_name,
                   r.name as reviewed_by_name
            FROM syllabus_uploads su
            JOIN users u ON su.uploaded_by = u.id
            LEFT JOIN departments d ON su.department_id = d.id
            LEFT JOIN users r ON su.reviewed_by = r.id
            ORDER BY su.uploaded_at DESC
        ''')
        
        uploads = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'uploads': [dict(row) for row in uploads]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/syllabus/approve/<int:upload_id>', methods=['POST'])
@jwt_required()
def approve_syllabus(upload_id):
    """Approve a syllabus upload"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        review_notes = data.get('review_notes', '')
        
        # Update syllabus upload status
        cursor.execute('''
            UPDATE syllabus_uploads 
            SET status = 'approved', reviewed_by = ?, review_notes = ?, reviewed_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (current_user_id, review_notes, upload_id))
        
        if cursor.rowcount == 0:
            return jsonify({'error': 'Upload not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Syllabus approved successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/syllabus/reject/<int:upload_id>', methods=['POST'])
@jwt_required()
def reject_syllabus(upload_id):
    """Reject a syllabus upload"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        review_notes = data.get('review_notes', '')
        
        # Update syllabus upload status
        cursor.execute('''
            UPDATE syllabus_uploads 
            SET status = 'rejected', reviewed_by = ?, review_notes = ?, reviewed_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (current_user_id, review_notes, upload_id))
        
        if cursor.rowcount == 0:
            return jsonify({'error': 'Upload not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Syllabus rejected successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/timetables/logs')
@jwt_required()
def timetable_logs():
    """View timetable generation logs"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get timetable logs
        cursor.execute('''
            SELECT tl.*, d.name as department_name, u.name as generated_by_name
            FROM timetable_logs tl
            JOIN departments d ON tl.department_id = d.id
            JOIN users u ON tl.generated_by = u.id
            ORDER BY tl.generated_at DESC
        ''')
        
        logs = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'logs': [dict(row) for row in logs]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/chatbot/query', methods=['POST'])
@jwt_required()
def chatbot_query():
    """Handle chatbot queries using Gemini AI"""
    try:
        current_user_id = get_jwt_identity()
        
        # Verify main admin access
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE id = ?', (current_user_id,))
        user_role = cursor.fetchone()
        
        if not user_role or user_role['role'] != 'main_admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        user_query = data.get('query', '').strip()
        
        if not user_query:
            return jsonify({'error': 'Query is required'}), 400
        
        # Simple rule-based responses for common queries
        responses = {
            'credentials': 'To generate credentials: Click "Generate Credentials" button. This will create usernames and passwords for staff and department admins who don\'t have credentials yet. Use "Export Credentials" to download them as Excel file.',
            'analytics': 'The Analytics section shows: Total departments, staff count, pending approvals, and timetable generations. These numbers update automatically as data changes in the system.',
            'notifications': 'To send notifications: Enter title and message, select recipient type (staff, dept_admin, or all), then click Send. Recent notifications are shown below the form.',
            'syllabus': 'Syllabus Review shows uploaded files. You can approve or reject each upload with optional review notes. Status changes are tracked with timestamps.',
            'timetables': 'Timetable Logs show all generated timetables with department name, generation type, creator, and timestamp. This helps track system usage.',
            'help': 'Available features: 1) Credential Generator 2) Analytics Summary 3) Notification Sender 4) Syllabus Review 5) Timetable Logs. Ask about any specific feature for detailed help.'
        }
        
        # Find best matching response
        query_lower = user_query.lower()
        response = responses.get('help')  # default
        
        for key, value in responses.items():
            if key in query_lower:
                response = value
                break
        
        return jsonify({
            'success': True,
            'response': response
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Helper function to log timetable generation
def log_timetable_generation(department_id, generation_type, generated_by, entries_count=0, status='completed'):
    """Log timetable generation activity"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO timetable_logs (department_id, generation_type, generated_by, entries_count, status)
            VALUES (?, ?, ?, ?, ?)
        ''', (department_id, generation_type, generated_by, entries_count, status))
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        print(f"Error logging timetable generation: {e}")